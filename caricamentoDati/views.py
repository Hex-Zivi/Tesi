from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import render, redirect
from django.db import transaction
from django.db.models import Count
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout

import csv
import datetime
import openpyxl
import re

from .models import *
from .forms import *

from django_auth_ldap.backend import LDAPBackend
import requests
from bs4 import BeautifulSoup


def is_admin(user):
    # Esempio: verifica se l'utente è autenticato e ha il ruolo di amministratore
    return user.is_authenticated and user.is_staff


# Aggiunge un decoratore @admin_required
def admin_required(view_func):
    decorated_view_func = user_passes_test(
        is_admin,
        login_url='/login/',  # URL della pagina di login
        redirect_field_name=None
    )(view_func)
    return decorated_view_func


# ===================== VALUTAZIONI / HOME ====================
@login_required
def valutazioni(request):
    context = {'valutazioni': Valutazione.objects.all().order_by('-anno'),
               'cf': request.user.codice_fiscale,
               'is_superuser': request.user.is_superuser,
               }
    return render(request, "caricamentoDati/valutazioni.html", context)


def crea_valutazione(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        anno = request.POST.get("anno")
        numero_pubblicazioni = request.POST.get("numero_pubblicazioni")

        nuova_valutazione = Valutazione(
            nome=nome, anno=anno, dataCaricamento=datetime.date.today(), status="Vuota", numeroPubblicazioni=numero_pubblicazioni)
        nuova_valutazione.save()

    return redirect('valutazioni')


def cancella_valutazione(request, valutazione_nome):
    valutazioneDaCancellate = Valutazione.objects.filter(nome=valutazione_nome)
    valutazioneDaCancellate.delete()
    return redirect('valutazioni')


@login_required
def chiudi_valutazione(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    valutazione.status = "Chiusa"
    valutazione.save()

    return redirect('valutazioni')


@login_required
def modifica_valutazione(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    form = FormAggiungiPubblicazione()

    context = {
        'valutazione': valutazione,
        'pubblicazioni': PubblicazionePresentata.objects.filter(valutazione=valutazione).order_by('titolo'),
        'form_aggiungi_pubblicazione': form,
    }
    return render(request, 'caricamentoDati/modifica.html', context)
# ===================== VALUTAZIONI / HOME ====================


# ===================== MODIFICA ====================
def caricamento_con_file(request, filename, valutazione):
    if request.method == 'POST':
        file = request.FILES.get('filename')
        valutazione = Valutazione.objects.get(nome=valutazione)
        elenco = []
        intestazione = ['anno_di_pubblicazione', 'autore', 'codice_fiscale', 'handle', 'doi', 'titolo',
                        'tipologia_collezione',
                        'issn_o_isbn', 'titolo_rivista_o_atti', 'indicizzato_scopus', 'miglior_quartile_scopus',
                        'num_coautori_interni_dip', 'codice_fiscale']
        riferimento = []
        if file:
            # Determina il tipo di file e leggi i dati appropriati
            if file.name.endswith('.csv'):
                csv_data = csv.reader(file.read().decode(
                    'utf-8').splitlines(), delimiter=',')
                for valore in csv_data:
                    elenco.append(valore)
            elif filename.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    elenco.append(row)
            else:
                # Gestisci il caso in cui il tipo di file non è supportato
                return HttpResponse("Il tipo di file non è supportato.")

            for element in elenco[5]:
                for titolo in intestazione:
                    if element.lower() == titolo:
                        riferimento.append(intestazione.index(titolo))

            with transaction.atomic():

                # In assenza di un file di caricamento veritiero, ho aggiunto una colonna per il codice fiscale alla fine: modificare i numeri delle colonne in maniera consona
                for riga in elenco[6:]:
                    anno_pubblicazione = riga[0]
                    autore = riga[1]
                    handle = riga[2]
                    doi = riga[3]
                    titolo = riga[4]
                    tipologia_collezione = riga[5]
                    issn_isbn = riga[6]
                    titolo_rivista_atti = riga[7]
                    titolo_rivista_atti = titolo_rivista_atti.upper()
                    indicizzato_scopus = riga[8].lower()
                    if indicizzato_scopus in ['vero', '1', 'true']:
                        indicizzato_scopus = True
                    else:
                        indicizzato_scopus = False
                    if riga[9] == '':
                        miglior_quartile = 0
                    else:
                        miglior_quartile = int(riga[9])
                    num_coautori_dip = riga[10]
                    codice_fiscale = riga[11]
                    autore = autore.upper()
                    codice_fiscale = codice_fiscale.upper()

                    if not Docente.objects.filter(codiceFiscale=codice_fiscale).exists():
                        Docente(codiceFiscale=codice_fiscale,
                                cognome_nome=autore).save()

                    if not PubblicazionePresentata.objects.filter(handle=handle).exists():
                        PubblicazionePresentata(handle=handle,
                                                issn_isbn=issn_isbn,
                                                anno_pubblicazione=anno_pubblicazione,
                                                doi=doi,
                                                titolo=titolo,
                                                tipologia_collezione=tipologia_collezione,
                                                titolo_rivista_atti=titolo_rivista_atti,
                                                indicizzato_scopus=indicizzato_scopus,
                                                miglior_quartile=miglior_quartile,
                                                num_coautori_dip=num_coautori_dip,
                                                valutazione=valutazione).save()

                    if not RelazioneDocentePubblicazione.objects.filter(pubblicazione=handle,
                                                                        autore__codiceFiscale=codice_fiscale).exists():
                        RelazioneDocentePubblicazione(pubblicazione=PubblicazionePresentata.objects.get(handle=handle),
                                                      autore=Docente.objects.get(codiceFiscale=codice_fiscale)).save()

        valutazione.status = "Pubblicazioni caricate"
        valutazione.dataCaricamentoPubblicazioni = datetime.date.today()
        valutazione.save()

    return redirect('modifica_valutazione', valutazione)


def cancella_pubblicazioni_tot(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    pubblicazioni = PubblicazionePresentata.objects.filter(
        valutazione=valutazione)
    for pubblicazione in pubblicazioni:
        pubblicazione.delete()
    return redirect('modifica_valutazione', valutazione)


def cancella_pubblicazione_singola(request, pubblicazione_titolo, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    PubblicazionePresentata.objects.get(titolo=pubblicazione_titolo).delete()
    return redirect('modifica_valutazione', valutazione)
# ===================== MODIFICA ====================


# ==================== AGGIUNTA PUBBLICAZIONE ====================
def aggiungi_pubblicazione_pagina(request, valutazione_nome, caller, docente):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    docente_chiamante = docente
    if docente != 'admin':
        docente = Docente.objects.get(codiceFiscale=docente)
        form = FormAggiungiPubblicazione(
            initial={'autori': [docente.codiceFiscale]})
    else:
        form = FormAggiungiPubblicazione()

    context = {
        'valutazione': valutazione,
        'caller': caller,
        'docente': docente,
        'form_aggiungi_pubblicazione': form,
        'docente_codice_fiscale': docente_chiamante,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string(
            'caricamentoDati/aggiungi_pubblicazione.html', context, request=request)
        return JsonResponse({'html': html})

    return render(request, 'caricamentoDati/aggiungi_pubblicazione.html', context)


def aggiungi_pubblicazione(request, valutazione_nome, docente, caller):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    docente_chiamante = ''
    if docente != 'admin':
        docente_chiamante = Docente.objects.get(codiceFiscale=docente)
        form = FormAggiungiPubblicazione(
            initial={'autori': [docente_chiamante.codiceFiscale]})
    else:
        form = FormAggiungiPubblicazione()

    if request.method == 'POST':
        form = FormAggiungiPubblicazione(request.POST)
        if form.is_valid():
            handle = form.cleaned_data['handle']
            titolo = form.cleaned_data['titolo']
            anno_pubblicazione = form.cleaned_data['anno_pubblicazione']
            miglior_quartile = form.cleaned_data['miglior_quartile']
            issn_isbn = form.cleaned_data['issn_isbn']
            doi = form.cleaned_data['doi']
            tipologia_collezione = form.cleaned_data['tipologia_collezione']
            titolo_rivista_atti = form.cleaned_data['titolo_rivista_atti']
            indicizzato_scopus = form.cleaned_data['indicizzato_scopus']
            numero_coautori_dip = len(form.cleaned_data['autori'])

            nuova_pubblicazione = PubblicazionePresentata(handle=handle, valutazione=valutazione, titolo=titolo, anno_pubblicazione=anno_pubblicazione, miglior_quartile=miglior_quartile,
                                                          issn_isbn=issn_isbn, doi=doi, tipologia_collezione=tipologia_collezione, titolo_rivista_atti=titolo_rivista_atti, indicizzato_scopus=indicizzato_scopus,
                                                          num_coautori_dip=numero_coautori_dip)
            nuova_pubblicazione.save()

            for docente in form.cleaned_data['autori']:
                nuova_relazione = RelazioneDocentePubblicazione(
                    autore=docente, pubblicazione=nuova_pubblicazione)
                nuova_relazione.save()

            if caller == "modifica":
                return redirect('modifica_valutazione', valutazione)

            elif caller == "assegnamento":
                return redirect('assegnamento', valutazione_nome)

            elif caller == "docente_pubblicazione":
                if docente_chiamante:
                    return redirect('docente_pubblicazioni', valutazione_nome, docente_chiamante.codiceFiscale)

            else:
                return redirect('valutazioni')

        else:
            form = FormAggiungiPubblicazione(request.POST)

    context = {
        'valutazione': valutazione,
        'docente': docente,
        'form_aggiungi_pubblicazione': form,
    }
    return render(request, 'caricamentoDati/aggiungi_pubblicazione.html', context)
# ==================== AGGIUNTA PUBBLICAZIONE ====================


# ================== RIVISTE ECCELLENTI ===================
def riviste_eccellenti(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    riviste = RivistaEccellente.objects.filter(valutazione=valutazione)
    context = {
        'valutazione': valutazione,
        'riviste': riviste
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string(
            'caricamentoDati/riviste_eccellenti.html', context, request=request)
        return JsonResponse({'html': html})

    return render(request, 'caricamentoDati/riviste_eccellenti.html', context)


def format_ISSN(issn_text):
    # Rimuove tutti i caratteri non numerici dall'ISSN
    digits = re.sub(r'\D', '', issn_text)
    # Formatta l'ISSN nel formato 'XXXX-XXXX'
    formatted_issn = '-'.join([digits[:4], digits[4:]])
    return formatted_issn


def carica_riviste(request, valutazione_nome):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('filename')
        valutazione = Valutazione.objects.get(nome=valutazione_nome)
        elenco = []
        intestazione = ['titolo', 'link']
        riferimento = []
        if xlsx_file:
            workbook = openpyxl.load_workbook(xlsx_file)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                elenco.append(row)

            with transaction.atomic():
                for riga in elenco[1:]:
                    nome = riga[0].upper()
                    link = riga[1]

                    response = requests.get(link)
                    soup = BeautifulSoup(response.content, 'html.parser')
                    issn_header = soup.find('h2', string="ISSN")
                    issn_element = issn_header.find_next(
                        'p') if issn_header else None
                    issn_text = issn_element.text.strip() if issn_element else "ISSN non trovato"

                    # Se ci sono più di un ISSN, dividi e formatta entrambi
                    issn_list = [format_ISSN(issn.strip())
                                 for issn in issn_text.split(',')]
                    issn1 = issn_list[0] if issn_list else ''
                    issn2 = issn_list[1] if len(issn_list) > 1 else ''

                    if not RivistaEccellente.objects.filter(valutazione=valutazione, nome=nome).exists():
                        RivistaEccellente.objects.create(
                            valutazione=valutazione,
                            nome=nome,
                            link=link,
                            issn1=issn1,
                            issn2=issn2
                        )

    return redirect('modifica_valutazione', valutazione)


def cancella_riviste(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    riviste = RivistaEccellente.objects.filter(
        valutazione=valutazione)
    for rivista in riviste:
        rivista.delete()
    return redirect('modifica_valutazione', valutazione)
# ================== RIVISTE ECCELLENTI ===================


# ===================== ASSEGNAMENTO ====================
@login_required
def assegnamento(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    relazioni_docente_pubblicazione = RelazioneDocentePubblicazione.objects.filter(
        pubblicazione__valutazione=valutazione)

    docenti_info = []

    for docente in Docente.objects.all().order_by('cognome_nome'):
        num_pubblicazioni_richieste = valutazione.numeroPubblicazioni
        num_pubblicazioni_assegnate = relazioni_docente_pubblicazione.filter(
            autore=docente).exclude(scelta__isnull=True).exclude(scelta=0).count()
        quartili = sorted(set(relazione.pubblicazione.miglior_quartile for relazione in relazioni_docente_pubblicazione.filter(
            autore=docente).exclude(scelta__isnull=True).exclude(scelta=0)), reverse=True)
        pubblicazioni_totali = relazioni_docente_pubblicazione.filter(
            autore=docente).count()
        formatted_quartili = sorted(['Q{}'.format(q) for q in quartili])
        condizione_pubblicazioni = 0
        if num_pubblicazioni_assegnate == num_pubblicazioni_richieste or (num_pubblicazioni_assegnate == pubblicazioni_totali and pubblicazioni_totali < num_pubblicazioni_richieste):
            condizione_pubblicazioni = 1

        docente_info = {
            'codice_fiscale': docente.codiceFiscale,
            'cognome_nome': docente.cognome_nome,
            'num_pubblicazioni_richieste': num_pubblicazioni_richieste,
            'num_pubblicazioni_assegnate': num_pubblicazioni_assegnate,
            'pubblicazioni_totali': pubblicazioni_totali,
            'quartili': formatted_quartili,
            'condizione_pubblicazioni': condizione_pubblicazioni

        }

        docenti_info.append(docente_info)

    informazioni = {'docenti_info': docenti_info}
    context = {'valutazione': valutazione,
               'informazioni_docente': informazioni}
    return render(request, 'caricamentoDati/assegnamento.html', context)


# Funzione di supporto per assegamento_algoritmo
def calcola_numero_coautori_possibili(pubblicazione, relazioni_docente_pubblicazione, docenti):
    return len(relazioni_docente_pubblicazione.filter(pubblicazione=pubblicazione))


# Funzione di supporto per assegamento_algoritmo
def rimuovi_docente(docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione):
    if docente in docenti:
        relazioni_docente_pubblicazione = relazioni_docente_pubblicazione.exclude(
            autore=docente)
        docenti = docenti.exclude(pk=docente.pk)
        for pubblicazione in pubblicazioni:
            numero_coautori_possibili_pubblicazione[pubblicazione.pk] = calcola_numero_coautori_possibili(
                pubblicazione, relazioni_docente_pubblicazione, docenti)
    return docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione


def assegnamento_algoritmo(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    numero_selezioni_valutazione = valutazione.numeroPubblicazioni
    pubblicazioni = PubblicazionePresentata.objects.filter(
        valutazione=valutazione).order_by('miglior_quartile')
    relazioni_docente_pubblicazione = RelazioneDocentePubblicazione.objects.filter(
        pubblicazione__valutazione=valutazione)
    docenti = Docente.objects.annotate(num_relazioni=Count(
        'relazionedocentepubblicazione')).order_by('num_relazioni')
    riviste_ecc = RivistaEccellente.objects.filter(valutazione=valutazione)

    numero_selezioni_docente = {docente.pk: 0 for docente in docenti}
    for docente in docenti:
        numero_selezioni_docente[docente.pk] = len(
            relazioni_docente_pubblicazione.filter(autore=docente, scelta=1))

    numero_coautori_possibili_pubblicazione = {}

    for pubblicazione in pubblicazioni:
        numero_coautori_possibili_pubblicazione[pubblicazione.pk] = calcola_numero_coautori_possibili(
            pubblicazione, relazioni_docente_pubblicazione, docenti)

    progresso = 1  # flag di verifica di aggiunta selezione

    for rivista in riviste_ecc:
        if rivista.issn2:
            lista_issn = rivista.issn1, rivista.issn2
        else:
            lista_issn = rivista.issn1

        for issn in lista_issn:
            pubblicazioni_rivista = pubblicazioni.filter(
                issn_isbn=issn)
            for pubblicazione in pubblicazioni_rivista:

                if pubblicazione.num_coautori_dip == 1:
                    relazione = RelazioneDocentePubblicazione.objects.get(
                        pubblicazione__handle=pubblicazione.handle)

                    docente = relazione.autore
                    if numero_selezioni_docente.get(docente.pk, 0) >= numero_selezioni_valutazione:
                        docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                            docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                        pubblicazioni = pubblicazioni.exclude(
                            relazionedocentepubblicazione__autore=docente, num_coautori_dip=1)
                        break

                    relazione.scelta = 1
                    relazione.save()

                    numero_selezioni_docente[docente.pk] += 1
                    if numero_selezioni_docente.get(docente.pk, 0) == numero_selezioni_valutazione:
                        docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                            docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                        progresso = 1
                        pubblicazioni = pubblicazioni.exclude(
                            relazionedocentepubblicazione__autore=docente, num_coautori_dip=1)

    while progresso == 1:
        journal_1_singoloAutore = pubblicazioni.filter(
            miglior_quartile__in=[0, 1], num_coautori_dip=1)
        journal_2_singoloAutore = pubblicazioni.filter(
            miglior_quartile=2, num_coautori_dip=1)
        journal_3_singoloAutore = pubblicazioni.filter(
            num_coautori_dip=1).exclude(miglior_quartile__in=[0, 1, 2])

        progresso = 0

        # Selezione di pubblicazioni con autore singolo
        for docente in docenti.order_by('cognome_nome'):
            for journal in [journal_1_singoloAutore, journal_2_singoloAutore, journal_3_singoloAutore]:
                if docente in docenti:
                    for pubblicazione in journal.filter(relazionedocentepubblicazione__autore=docente):

                        if numero_selezioni_docente.get(docente.pk, 0) >= numero_selezioni_valutazione:
                            docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                                docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                            pubblicazioni = pubblicazioni.exclude(
                                relazionedocentepubblicazione__autore=docente, num_coautori_dip=1)
                            break

                        if pubblicazione in pubblicazioni.filter(relazionedocentepubblicazione__autore=docente) and relazioni_docente_pubblicazione.get(autore=docente, pubblicazione=pubblicazione).scelta == 0:
                            relazione = RelazioneDocentePubblicazione.objects.get(
                                pubblicazione=pubblicazione, autore=docente)
                            relazione.scelta = 1
                            relazione.save()

                            progresso = 1

                            numero_selezioni_docente[docente.pk] += 1
                            if numero_selezioni_docente.get(docente.pk, 0) == numero_selezioni_valutazione:
                                docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                                    docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                                progresso = 1
                                pubblicazioni = pubblicazioni.exclude(
                                    relazionedocentepubblicazione__autore=docente, num_coautori_dip=1)
                                break

        # Selezione di pubblicazioni con autori multipli
        if progresso == 0:
            for pubblicazione in pubblicazioni:
                numero_coautori_possibili_pubblicazione[pubblicazione.pk] = calcola_numero_coautori_possibili(
                    pubblicazione, relazioni_docente_pubblicazione, docenti)

            journal_1 = pubblicazioni.filter(
                miglior_quartile__in=[0, 1]).exclude(num_coautori_dip=1)
            journal_2 = pubblicazioni.filter(
                miglior_quartile=2).exclude(num_coautori_dip=1)
            journal_3 = pubblicazioni.exclude(
                miglior_quartile__in=[0, 1, 2], num_coautori_dip=1)

            for journal in [journal_1, journal_2, journal_3]:
                for docente in docenti:

                    for pubblicazione in journal.filter(relazionedocentepubblicazione__autore=docente):

                        if numero_selezioni_docente.get(docente.pk, 0) >= numero_selezioni_valutazione:
                            docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                                docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                            break

                        if numero_coautori_possibili_pubblicazione[pubblicazione.pk] <= 1 and pubblicazione in pubblicazioni.filter(relazionedocentepubblicazione__autore=docente) and relazioni_docente_pubblicazione.get(autore=docente, pubblicazione=pubblicazione).scelta == 0:
                            relazione = RelazioneDocentePubblicazione.objects.get(
                                pubblicazione=pubblicazione, autore=docente)
                            relazione.scelta = 1
                            relazione.save()

                            progresso = 1

                            numero_selezioni_docente[docente.pk] += 1
                            if numero_selezioni_docente.get(docente.pk, 0) == numero_selezioni_valutazione:
                                docenti, numero_coautori_possibili_pubblicazione, relazioni_docente_pubblicazione = rimuovi_docente(
                                    docente, pubblicazioni, relazioni_docente_pubblicazione, docenti, numero_coautori_possibili_pubblicazione)
                                progresso = 1
                                break

    valutazione.status = "Assegnamento calcolato"
    valutazione.save()

    return redirect('assegnamento', valutazione)


def azzera_assegnamento(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    relazioni_docente_pubblicazione = RelazioneDocentePubblicazione.objects.filter(
        pubblicazione__valutazione=valutazione)

    for relazione in relazioni_docente_pubblicazione:
        relazione.scelta = 0
        relazione.save()

    valutazione.status = "Pubblicazioni caricate"
    valutazione.save()

    return redirect('assegnamento', valutazione)
# ===================== ASSEGNAMENTO ====================


# =================== DOCENTE-PUBBLICAZIONI ====================
def docente_pubblicazioni(request, valutazione_nome, docente_codice_fiscale):
    docente_cf = request.session.get('codice_fiscale')
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    try:
        docente = Docente.objects.get(codiceFiscale=docente_cf)
    except Docente.DoesNotExist:
        docente_nome = str(request.session.get('last_name')).upper(
        ) + " " + str(request.session.get('first_name')).upper()
        docente = Docente(cognome_nome=docente_nome, codiceFiscale=docente_cf)
        docente.save()

    relazioni_docente_pubblicazione = RelazioneDocentePubblicazione.objects.filter(
        pubblicazione__valutazione=valutazione, autore=docente)

    pubblicazioni_info = []

    for relazione in relazioni_docente_pubblicazione:
        pubblicazione = relazione.pubblicazione
        altri_autori = [autore.cognome_nome for autore in Docente.objects.filter(
            relazionedocentepubblicazione__pubblicazione=pubblicazione
        ).exclude(codiceFiscale=docente_cf).distinct()]
        altri_autori_scelta = [autore.cognome_nome for autore in Docente.objects.filter(
            relazionedocentepubblicazione__pubblicazione=pubblicazione, relazionedocentepubblicazione__scelta__gt=0
        ).exclude(codiceFiscale=docente_cf).distinct()]
        valore_scelta = relazione.scelta
        quartile = pubblicazione.miglior_quartile

        pubblicazione_info = {
            'titolo': pubblicazione.titolo,
            'anno': pubblicazione.anno_pubblicazione,
            'tipologia': pubblicazione.tipologia_collezione,
            'rivista': pubblicazione.titolo_rivista_atti,
            'altri_autori': altri_autori,
            'altri_autori_scelta': altri_autori_scelta,
            'valore_scelta': valore_scelta,
            'quartile': quartile,
            'doi': pubblicazione.doi,
            'issn_isbn': pubblicazione.issn_isbn,
            'handle': pubblicazione.handle,
            'scopus': pubblicazione.indicizzato_scopus,
        }

        pubblicazioni_info.append(pubblicazione_info)

    context = {'valutazione': valutazione,
               'num_pubblicazioni': valutazione.numeroPubblicazioni,
               'docente': docente,
               'docente_codice_fiscale': docente.codiceFiscale,
               'pubblicazioni_info': pubblicazioni_info}

    return render(request, 'caricamentoDati/docente_pubblicazioni.html', context)


def salva_selezioni(request, valutazione_nome, docente_codice_fiscale):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    docente = Docente.objects.get(codiceFiscale=docente_codice_fiscale)

    if request.method == 'POST':
        titoli_pubblicazioni_selezionate = request.POST.getlist(
            'titoli_pubblicazione[]')

        relazioni_docente_pubblicazione = RelazioneDocentePubblicazione.objects.filter(
            pubblicazione__valutazione=valutazione, autore=docente)

        for relazione in relazioni_docente_pubblicazione:
            relazione.scelta = 0
            relazione.save()

        for titolo_pubblicazione in titoli_pubblicazioni_selezionate:
            relazione = RelazioneDocentePubblicazione.objects.get(
                pubblicazione__titolo=titolo_pubblicazione,
                autore__codiceFiscale=docente_codice_fiscale
            )
            if relazione:
                relazione.scelta = 1
                relazione.save()

    if is_admin(request.user):
        return redirect('assegnamento', valutazione_nome)

    else:
        return redirect('valutazioni')
# =================== DOCENTE-PUBBLICAZIONI ====================


# =================== SELEZIONI ====================
def visualizza_selezioni(request, valutazione_nome):
    valutazione = Valutazione.objects.get(nome=valutazione_nome)
    selezioni_per_autore = {}

    docenti = Docente.objects.all().order_by('cognome_nome')

    for docente in docenti:
        pubblicazioni_selezionate = RelazioneDocentePubblicazione.objects.filter(
            autore=docente, pubblicazione__valutazione=valutazione, scelta=1).select_related('pubblicazione')
        selezioni_per_autore[docente] = pubblicazioni_selezionate

    context = {
        'valutazione': valutazione,
        'selezioni_per_autore': selezioni_per_autore
    }

    return render(request, 'caricamentoDati/selezioni.html', context)
# =================== SELEZIONI ====================
